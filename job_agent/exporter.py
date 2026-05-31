"""
Exports filtered job listings to a formatted Excel (.xlsx) file.
"""
import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Columns exported and their display widths (chars)
COLUMNS = [
    ("Job Role",     30),
    ("Company",      28),
    ("Location",     22),
    ("Posted",       18),
    ("Company Size", 16),
    ("Industry",     28),
    ("Source",       12),
    ("Job URL",      55),
]

HEADER_BG   = "1F3864"   # dark navy
TITLE_BG    = "2E75B6"   # medium blue
ROW_ALT_BG  = "EBF3FB"   # light blue tint
URL_COLOR   = "0563C1"


def export(jobs: list[dict], output_dir: str) -> str:
    """Write jobs to a timestamped Excel file. Returns the file path."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    stamp    = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filepath = output_dir / f"jobs_{stamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Results"

    n_cols = len(COLUMNS)
    last_col = get_column_letter(n_cols)

    # ------------------------------------------------------------------
    # Row 1 — banner title
    # ------------------------------------------------------------------
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = (
        f"Job Search Results  —  {datetime.now().strftime('%d %b %Y, %H:%M')}"
    )
    title_cell.font      = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ------------------------------------------------------------------
    # Row 2 — summary count
    # ------------------------------------------------------------------
    ws.merge_cells(f"A2:{last_col}2")
    summary = ws["A2"]
    summary.value     = f"Total matching jobs found: {len(jobs)}"
    summary.font      = Font(bold=True, size=11, color=HEADER_BG)
    summary.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    # ------------------------------------------------------------------
    # Row 3 — column headers
    # ------------------------------------------------------------------
    HEADER_ROW = 3
    thin = Side(border_style="thin", color="FFFFFF")
    for col_idx, (header, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=TITLE_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = Border(bottom=thin)
    ws.row_dimensions[HEADER_ROW].height = 24

    # ------------------------------------------------------------------
    # Data rows
    # ------------------------------------------------------------------
    field_map = {
        "Job Role":     "title",
        "Company":      "company",
        "Location":     "location",
        "Posted":       "posted_at",
        "Company Size": "company_size",
        "Industry":     "industry",
        "Source":       "source",
        "Job URL":      "job_url",
    }

    for row_offset, job in enumerate(jobs):
        row_num = HEADER_ROW + 1 + row_offset
        bg = ROW_ALT_BG if row_offset % 2 == 0 else "FFFFFF"

        for col_idx, (header, _) in enumerate(COLUMNS, start=1):
            field = field_map[header]
            value = job.get(field, "") or ""

            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="center", wrap_text=(header == "Job URL"))

            if header == "Job URL" and value:
                cell.hyperlink = value
                cell.font = Font(color=URL_COLOR, underline="single")

        ws.row_dimensions[row_num].height = 20

    # Empty-state message
    if not jobs:
        data_row = HEADER_ROW + 1
        ws.merge_cells(f"A{data_row}:{last_col}{data_row}")
        cell = ws.cell(row=data_row, column=1,
                       value="No jobs found matching all criteria today.")
        cell.font      = Font(italic=True, color="888888", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[data_row].height = 24

    # ------------------------------------------------------------------
    # Column widths, freeze panes, auto-filter
    # ------------------------------------------------------------------
    for col_idx, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)
    ws.auto_filter.ref = (
        f"A{HEADER_ROW}:{last_col}{max(HEADER_ROW + len(jobs), HEADER_ROW + 1)}"
    )

    wb.save(filepath)
    logger.info(f"[Exporter] Saved -> {filepath}")
    return str(filepath)
